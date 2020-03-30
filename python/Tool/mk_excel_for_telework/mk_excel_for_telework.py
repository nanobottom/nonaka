import xlsxwriter
import datetime
import locale
import calendar
import jholiday
import numpy as np
import os
import csv
import configparser
from openpyxl.utils import get_column_letter
from editcell import EditCell

# configparserの宣言とiniファイルの読み込み
config_ini = configparser.ConfigParser()
config_ini.read('config.ini', encoding='utf-8')

# 設定ファイルの値をセットする
current_year = config_ini.getint('DEFAULT', 'start_year')
current_month = config_ini.getint('DEFAULT', 'start_month')
fin_month_interval = config_ini.getint('DEFAULT', 'end_month_interval')
main_column_num = 2
filename = config_ini.get('DEFAULT', 'save_filename')
name_filename = config_ini.get('DEFAULT', 'name_csv_filename')

# 名前のリスト（CSVファイル）を読み込む
names = list()
with open(name_filename, 'r') as csv_file:
    f = csv.reader(csv_file)
    for i in f:
        names.append(i)
names = list(np.array(names).flatten()) # numpyのflattenを使用して平滑化

white_row_num = len(names)

"""
 土日祝日を判定して土曜日なら青,祝日か日曜日なら赤を、
 それ以外ならcolorを返す関数
"""
def is_sat_sun_holiday(date, color):
    day_of_the_week_index = int(date.strftime('%w'))
    if jholiday.holiday_name(date.year, date.month, date.day) is not None:
        return 'red'
    else:
        if day_of_the_week_index == 6:
            return 'blue'
        elif day_of_the_week_index == 0:
            return 'red'
        else:
            return color

# 現在のディレクトリから、指定した名前のxlsxファイルの有無を調べる。
# 既にある場合はエラーを出力し、なければシートを新規で作成する
current_dir = os.path.dirname(os.path.abspath(__file__))
xlsx_path = os.path.join(current_dir, filename)
try:
    if os.path.exists(xlsx_path) == True:
        raise SystemError('{}の"{}"が上書きされる恐れがあります。'.format(current_dir, filename))
except SystemError as e:
    print("SystemError : {}".format(e))
    exit()
else:
    wb = xlsxwriter.Workbook(filename)
sheet_name = str(current_year) + '年' + str(current_month) + '月～'

ws = wb.add_worksheet(sheet_name)

# 名前(A1)のセル編集
ws.merge_range("A1:A3", 'Merged Range')
a1 = EditCell(wb, ws, 1, 1, '名前')
a1.edit_font(bold = True)
a1.edit_border_left()
a1.edit_border_top()
a1.edit_border_right()
a1.edit_height_width(width = 10)
a1.edit_fill(fgColor = 'cyan')
a1.edit_alignment()

a2 = EditCell(wb, ws, 1, 2)
a2.edit_border_left()
a2.edit_border_right()

a3 = EditCell(wb, ws, 1, 3)
a3.edit_border_left()
a3.edit_border_right()
a3.edit_border_bottom()

# 備考(B1)のセル編集
ws.merge_range("B1:B3", 'Merged Range')
b1 = EditCell(wb, ws, 2, 1, '備考')
b1.edit_font(bold = True)
b1.edit_border_left()
b1.edit_border_top()
b1.edit_border_right()
b1.edit_height_width(width = 30)
b1.edit_fill(fgColor = 'cyan')
b1.edit_alignment()

b2 = EditCell(wb, ws, 2, 2)
b2.edit_border_left()
b2.edit_border_right()

b3 = EditCell(wb, ws, 2, 3)
b3.edit_border_left()
b3.edit_border_right()
b3.edit_border_bottom()

#白枠（4行目以降）の編集
for i_row in range(4, white_row_num + 4):
    for i_column in range(1, main_column_num + 1):
        # 1列目にリストから抽出した名前を書き込む
        if i_column == 1:
            white = EditCell(wb, ws, i_column, i_row, names[i_row - 4])
        else:
            white = EditCell(wb, ws, i_column, i_row)
        white.edit_fill(fgColor = 'white')
        white.edit_alignment()
        white.edit_border_round()

# 列の固定
ws.freeze_panes(0, main_column_num) 

date = datetime.datetime(current_year, current_month, 1)
column_count = main_column_num + 1
for i_month in range(current_month, current_month + fin_month_interval):
    move_up_year = i_month // 12
    if move_up_year > 0 and i_month % 12 != 0:
        i_month -= move_up_year * 12
    (_, last_day) = calendar.monthrange(current_year + move_up_year, i_month)

    for i_day in range(1,last_day + 1):
        # 年月日を取得する
        year = date.year
        month = date.month
        day = date.day

        # 日付の文字列からその日付の曜日を取得する
        locale.setlocale(locale.LC_TIME, 'ja_JP.UTF-8')
        day_of_the_week = date.strftime('%a')
        day_of_the_week_index = int(date.strftime('%w'))
        
        # 日付セルの編集
        if i_day == 1:
            # 年度の最初の月はxxxx年xx月の文字列を表示させるためにセルを結合する
            if i_month == 1:
                month_title = str(year) + '年' + str(month) + '月'
                ws.merge_range(get_column_letter(column_count) + '1:' + get_column_letter(column_count + 2) + '1', 'Merged Range')
            # 年度の最初の月以外はxx月と表示させる
            else:
                ws.merge_range(get_column_letter(column_count) + '1:' + get_column_letter(column_count + 1) + '1', 'Merged Range')
                month_title = str(month) + '月'
            i1= EditCell(wb, ws, column_count, 1, month_title)
            i1.edit_border_left()
        else:
            i1= EditCell(wb, ws, column_count, 1)
        i1.edit_font(bold = True)
        i1.edit_fill(fgColor = 'orange')
        i1.edit_height_width(width = 4)
        i1.edit_alignment()

        #土曜日なら青、日曜日、祝日なら赤、それ以外は黄色にして日付と曜日を表示させる
        ans_color = is_sat_sun_holiday(date, 'yellow')
        i2= EditCell(wb, ws, column_count, 2, str(day))
        i2.edit_font(bold = True)
        i2.edit_fill(fgColor = ans_color)
        i2.edit_height_width(width = 4)
        i2.edit_alignment()
        i2.edit_border_round()

        i3 = EditCell(wb, ws, column_count, 3, day_of_the_week)
        i3.edit_font(bold = True)
        i3.edit_fill(fgColor = ans_color)
        i3.edit_height_width(width = 4)
        i3.edit_alignment()
        i3.edit_border_round()

        #白枠（4行目以降）の編集(土日祝で色を変更する)
        for i_row in range(4, white_row_num + 4):
            ans_color = is_sat_sun_holiday(date, 'white')
            white_cell = EditCell(wb, ws, column_count, i_row)
            white_cell.edit_fill(fgColor = ans_color)
            white_cell.edit_height_width(width = 4)
            white_cell.edit_alignment()
            white_cell.edit_border_round()
            white_cell.regulate_input_str(['○', '　'])
        print('Processing : {0}年{1}月{2}日({3})'.format(year, month, day, day_of_the_week))
        date = date + datetime.timedelta(days = 1)
        column_count += 1
    current_month += 1
wb.close()
print('Finish!')
