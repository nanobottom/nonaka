import xlsxwriter, openpyxl
from openpyxl.utils import get_column_letter
class EditCell:

    def __init__(self, workbook, worksheet, column, row, value = ''):
        self.wb = workbook
        self.ws = worksheet
        self.column = column - 1
        self.row = row - 1
        self.cell = get_column_letter(column) + str(row)
        self.fmt = workbook.add_format( )
        self.value = value
    # underline:1=single underline, 2 = double underline, 33=single accounting underline, 34 = double accounting underline
    def edit_font(self, name = 'ＭＳ Ｐゴシック', size = 11, bold = False, italic = False, underline = 'none', strike = False, color = '#000000'):
        self.fmt.set_font_name(name)
        self.fmt.set_font_size(size)
        self.fmt.set_font_color(color)
        if bold is True:
            self.fmt.set_bold()
        if italic is True:
            self.fmt.set_italic()
        if underline != 'none':
            self.fmt.set_underline(underline)
        if strike is True:
            self.fmt.set_font_strikeout()

    def edit_border_round(self, border_style = 1, border_color = '#000000'):
        self.fmt.set_border(border_style)
        self.fmt.set_border_color(border_color)
        self.ws.write(self.row, self.column, self.value, self.fmt)

    def edit_border_top(self, border_style = 1, border_color = '#000000'):
        self.fmt.set_top(border_style)
        self.fmt.set_top_color(border_color)
        self.ws.write(self.row, self.column, self.value, self.fmt)

    def edit_border_left(self, border_style = 1, border_color = '#000000'):
        self.fmt.set_left(border_style)
        self.fmt.set_left_color(border_color)
        self.ws.write(self.row, self.column, self.value, self.fmt)

    def edit_border_right(self, border_style = 1, border_color = '#000000'):
        self.fmt.set_right(border_style)
        self.fmt.set_right_color(border_color)
        self.ws.write(self.row, self.column, self.value, self.fmt)

    def edit_border_bottom(self, border_style = 1, border_color = '#000000'):
        self.fmt.set_bottom(border_style)
        self.fmt.set_bottom_color(border_color)
        self.ws.write(self.row, self.column, self.value, self.fmt)

    def edit_alignment(self, horizontal = 'center', vertical = 'vcenter'):
        self.fmt.set_align(horizontal)
        self.fmt.set_align(vertical)
        self.ws.write(self.row, self.column, self.value, self.fmt)

    def edit_fill(self, fill_type = 1, fgColor = '#FFFFFF'):
        self.fmt.set_pattern(fill_type)
        self.fmt.set_bg_color(fgColor)
        self.ws.write(self.row, self.column, self.value, self.fmt)

    def edit_height_width(self, width = 8.11, height = 13.2):
        self.ws.set_row(self.row, height)
        self.ws.set_column(self.column,self.column, width)

    def regulate_input_str(self, list_str):
        self.ws.data_validation(self.cell, {'validate': 'list', 'source': list_str})

    # criteria is 'between' or 'not between'
    def regulate_value_range(self, minimum, maximum, criteria = 'between'):
        self.ws.data_validation(self.cell, {'validate': 'decimal', 'criteria': criteria, 'minimum':minimum, 'maximum': maximum})

if __name__ == '__main__':
    import os
    sheet_name = 'test'
    filename = 'example.xlsx'
    current_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(current_dir, filename)
    try:
        if os.path.exists(xlsx_path) == True:
            raise SystemError('{}の"{}"が上書きされる恐れがあります。'.format(current_dir, filename))
    except SystemError as e:
        print("SystemError : {}".format(e))
        exit()
    else:
        workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()
    b2 = EditCell(workbook, worksheet, 2, 2, '項目')
    b2.edit_font(bold = True)
    b2.edit_border_right()
    b2.edit_border_left()
    b2.edit_alignment()
    b2.edit_fill(fgColor = '#FFFF00')
    b2.edit_height_width(2, 2)
    workbook.close()
