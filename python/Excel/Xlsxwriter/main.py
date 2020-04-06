import xlsxwriter, datetime, locale, calendar, jholiday, os
from openpyxl.utils import get_column_letter
from editcell import EditCell
current_year = 2020
current_month = 4
fin_month_interval = 1
white_row_num = 30
main_column_num = 8
project_num = 3
filename = '進捗管理表.xlsx'


# 土日祝日を判定して土曜日なら青,祝日か日曜日なら赤を、それ以外ならcolorを返す関数
def is_sat_sun_holiday(date, color):
    day_of_the_week_index = int(date.strftime('%w'))
    if jholiday.holiday_name(date.year, date.month, date.day) is not None:
        return 'red'
    else:
        if day_of_the_week_index == 6:
            return 'blue'
        elif day_of_the_week_index == 0:
            return 'red'
        else:
            return color
current_dir = os.path.dirname(os.path.abspath(__file__))
xlsx_path = os.path.join(current_dir, filename)
try:
    if os.path.exists(xlsx_path) == True:
        raise SystemError('{}の"{}"が上書きされる恐れがあります。'.format(current_dir, filename))
except SystemError as e:
    print("SystemError : {}".format(e))
    exit()
else:
    wb = xlsxwriter.Workbook(filename)
sheet_name = str(current_year) + '年' + str(current_month) + '月'
ws = wb.add_worksheet(sheet_name)

# 番号(A1)のセル編集
ws.merge_range("A1:A3", 'Merged Range')
a1 = EditCell(wb, ws, 1, 1, '番号')
a1.edit_font(bold = True)
a1.edit_border_left()
a1.edit_border_top()
a1.edit_border_right()
a1.edit_height_width(width = 6)
a1.edit_fill(fgColor = 'cyan')
a1.edit_alignment()

a2 = EditCell(wb, ws, 1, 2)
a2.edit_border_left()
a2.edit_border_right()

a3 = EditCell(wb, ws, 1, 3)
a3.edit_border_left()
a3.edit_border_right()
a3.edit_border_bottom()

# 項目(B1)のセル編集
ws.merge_range("B1:B3", 'Merged Range')
b1 = EditCell(wb, ws, 2, 1, '項目')
b1.edit_font(bold = True)
b1.edit_border_left()
b1.edit_border_top()
b1.edit_border_right()
b1.edit_height_width(width = 30)
b1.edit_fill(fgColor = 'cyan')
b1.edit_alignment()

b2 = EditCell(wb, ws, 2, 2)
b2.edit_border_left()
b2.edit_border_right()

b3 = EditCell(wb, ws, 2, 3)
b3.edit_border_left()
b3.edit_border_right()
b3.edit_border_bottom()
# 予定(C1)のセル編集
ws.merge_range("C1:D2", 'Merged Range')
c1 = EditCell(wb, ws, 3, 1, '予定')
c1.edit_font(bold = True)
c1.edit_border_left()
c1.edit_border_top()
c1.edit_fill(fgColor = 'cyan')
c1.edit_alignment()

c2 = EditCell(wb, ws, 3, 2)
c2.edit_border_left()
c2.edit_border_bottom()

d1 = EditCell(wb, ws, 4, 1)
d1.edit_border_right()
d1.edit_border_top()

d2 = EditCell(wb, ws, 4, 2)
d2.edit_border_right()
d2.edit_border_bottom()

# 実績(E1)のセル編集
ws.merge_range("E1:F2", 'Merged Range')
e1 = EditCell(wb, ws, 5, 1, '実績')
e1.edit_font(bold = True)
e1.edit_border_left()
e1.edit_border_top()
e1.edit_fill(fgColor = 'cyan')
e1.edit_alignment()

e2 = EditCell(wb, ws, 5, 2)
e2.edit_border_left()
e2.edit_border_bottom()

f1 = EditCell(wb, ws, 6, 1)
f1.edit_border_right()
f1.edit_border_top()

f2 = EditCell(wb, ws, 6, 2)
f2.edit_border_right()
f2.edit_border_bottom()

#開始(C3)のセル編集
c3 = EditCell(wb, ws, 3, 3, '開始')
c3.edit_font(bold = True)
c3.edit_fill(fgColor = 'cyan')
c3.edit_height_width(width = 6)
c3.edit_alignment()
c3.edit_border_round()

#終了(D3)のセル編集
d3 = EditCell(wb, ws, 4, 3, '終了')
d3.edit_font(bold = True)
d3.edit_fill(fgColor = 'cyan')
d3.edit_height_width(width = 6)
d3.edit_alignment()
d3.edit_border_round()

#開始(E3)のセル編集
e3 = EditCell(wb, ws, 5, 3, '開始')
e3.edit_font(bold = True)
e3.edit_fill(fgColor = 'cyan')
e3.edit_height_width(width = 6)
e3.edit_alignment()
e3.edit_border_round()

#終了(F3)のセル編集
f3 = EditCell(wb, ws, 6, 3, '終了')
f3.edit_font(bold = True)
f3.edit_fill(fgColor = 'cyan')
f3.edit_height_width(width = 6)
f3.edit_alignment()
f3.edit_border_round()

#工数(G1)のセル編集
ws.merge_range("G1:G3", 'Merged Range')
g1 = EditCell(wb, ws, 7, 1, '工数')
g1.edit_font(bold = True)
g1.edit_border_left()
g1.edit_border_top()
g1.edit_border_right()
g1.edit_height_width(width = 6)
g1.edit_fill(fgColor = 'cyan')
g1.edit_alignment()

g2 = EditCell(wb, ws, 7, 2)
g2.edit_border_left()
g2.edit_border_right()

g3 = EditCell(wb, ws, 7, 3)
g3.edit_border_left()
g3.edit_border_right()
g3.edit_border_bottom()

#状態(H1)のセル編集
ws.merge_range("H1:H3", 'Merged Range')
h1 = EditCell(wb, ws, 8, 1, '状態')
h1.edit_font(bold = True)
h1.edit_border_left()
h1.edit_border_top()
h1.edit_border_right()
h1.edit_fill(fgColor = 'cyan')
h1.edit_alignment()

h2 = EditCell(wb, ws, 8, 2)
h2.edit_border_left()
h2.edit_border_right()

h3 = EditCell(wb, ws, 8, 3)
h3.edit_border_left()
h3.edit_border_right()
h3.edit_border_bottom()

#白枠（4行目以降）の編集
for i_row in range(4, white_row_num + 4):
    for i_column in range(1, main_column_num + 1):
        white = EditCell(wb, ws, i_column, i_row)
        white.edit_fill(fgColor = 'white')
        white.edit_alignment()
        white.edit_border_round()

# 列の固定
ws.freeze_panes(0, main_column_num) 

date = datetime.datetime(current_year, current_month, 1)
column_count = main_column_num + 1
total_days = 0
for i_month in range(current_month, current_month + fin_month_interval):
    move_up_year = i_month // 12
    if move_up_year > 0 and i_month % 12 != 0:
        i_month -= move_up_year * 12
    (_, last_day) = calendar.monthrange(current_year + move_up_year, i_month)

    for i_day in range(1,last_day + 1):
        total_days += 1
        # 年月日を取得する
        year = date.year
        month = date.month
        day = date.day

        # 日付の文字列からその日付の曜日を取得する
        locale.setlocale(locale.LC_TIME, 'ja_JP.UTF-8')
        day_of_the_week = date.strftime('%a')
        day_of_the_week_index = int(date.strftime('%w'))
        
        # 日付セルの編集
        if i_day == 1:
            # 年度の最初の月はxxxx年xx月の文字列を表示させるためにセルを結合する
            if i_month == 1:
                month_title = str(year) + '年' + str(month) + '月'
                ws.merge_range(get_column_letter(column_count) + '1:' + get_column_letter(column_count + 2) + '1', 'Merged Range')
            # 年度の最初の月以外はxx月と表示させる
            else:
                ws.merge_range(get_column_letter(column_count) + '1:' + get_column_letter(column_count + 1) + '1', 'Merged Range')
                month_title = str(month) + '月'
            i1= EditCell(wb, ws, column_count, 1, month_title)
            i1.edit_border_left()
        else:
            i1= EditCell(wb, ws, column_count, 1)
        i1.edit_font(bold = True)
        i1.edit_fill(fgColor = 'orange')
        i1.edit_height_width(width = 4)
        i1.edit_alignment()

        #土曜日なら青、日曜日、祝日なら赤、それ以外は黄色にして日付と曜日を表示させる
        ans_color = is_sat_sun_holiday(date, 'yellow')
        i2= EditCell(wb, ws, column_count, 2, str(day))
        i2.edit_font(bold = True)
        i2.edit_fill(fgColor = ans_color)
        i2.edit_height_width(width = 4)
        i2.edit_alignment()
        i2.edit_border_round()

        i3 = EditCell(wb, ws, column_count, 3, day_of_the_week)
        i3.edit_font(bold = True)
        i3.edit_fill(fgColor = ans_color)
        i3.edit_height_width(width = 4)
        i3.edit_alignment()
        i3.edit_border_round()

        #白枠（4行目以降）の編集(土日祝で色を変更する)
        for i_row in range(4, white_row_num + 4):
            ans_color = is_sat_sun_holiday(date, 'white')
            white_cell = EditCell(wb, ws, column_count, i_row)
            white_cell.edit_fill(fgColor = ans_color)
            white_cell.edit_height_width(width = 4)
            white_cell.edit_alignment()
            white_cell.edit_border_round()
        print('Processing : {0}年{1}月{2}日({3})'.format(year, month, day, day_of_the_week))
        date = date + datetime.timedelta(days = 1)
        column_count += 1
    current_month += 1
# 工数の数式（全日程の時間の合計を算出する）を代入する
for i_row in range(4,white_row_num + 4):
    numerical_formula_sum ='=SUM(I'+str(i_row)+':'+get_column_letter(column_count - 1)+str(i_row)+')'
    ws.write(i_row - 1, 6, numerical_formula_sum, white.fmt)



# プロジェクト毎の工数を算出する
for i_project in range(1, project_num + 1):
    numerical_formula_sumif = '=SUMIF(A4:A' + str(white_row_num + 4) + ',' + str(i_project) + ',G4:G' + str(white_row_num + 4) + ')'
    p_title_cell = EditCell(wb, ws, i_project + 2, white_row_num + 6, 'P' + str(i_project) )
    p_title_cell.edit_font(bold = True)
    p_title_cell.edit_fill(fgColor = 'cyan')
    p_title_cell.edit_alignment()
    p_title_cell.edit_border_round()

    nf_cell = EditCell(wb, ws, i_project + 2, white_row_num + 7, numerical_formula_sumif)
    nf_cell.edit_font(bold = True)
    nf_cell.edit_fill(fgColor = 'white')
    nf_cell.edit_alignment()
    nf_cell.edit_border_round()

# 状態のプルダウンリストを作成する
for i_row in range(4,white_row_num + 4):
    status_cell = EditCell(wb, ws, 8, i_row)
    status_cell.edit_fill(fgColor = 'white')
    status_cell.edit_alignment()
    status_cell.edit_border_round()
    status_cell.regulate_input_str(['未着手', '進行中', '完了'])
# 1日毎の工数の合計を算出する
total_cell = EditCell(wb, ws, 2, 4, value='合計')
total_cell.edit_alignment()
numerical_formula = '=SUM(G5:G'+str(4 + white_row_num - 1)+')'
kousu_cell = EditCell(wb, ws, 7, 4, numerical_formula)
kousu_cell.edit_alignment()

for i_column in range(9, total_days + 9):
    numerical_formula = '=SUM(' + get_column_letter(i_column) + '5:' + get_column_letter(i_column) + str(4 + white_row_num - 1) + ')'
    current_cell = EditCell(wb, ws, i_column, 4, numerical_formula)
    current_cell.edit_border_round()

wb.close()
print('Finish!')
