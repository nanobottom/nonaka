import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
class EditCell:

    def __init__(self, work_sheet, column, row, value = ''):
        self.ws = work_sheet
        self.cell = self.ws.cell(row = row, column = column)
        self.column = column
        self.row = row
        if value != '':
            self.cell.value = value

    # strike:打ち消し線
    def edit_font(self, name = 'ＭＳ Ｐゴシック', size = 11, bold = False, italic = False, underline = 'none', strike = False, color = 'FF000000'):
        self.cell.font = Font(name = name, size = size, bold = bold, italic = italic, underline = underline, strike = strike, color = color)

    # セルの枠線を一周設定する
    def edit_border_round(self, border_style = 'thin', border_color = 'FF000000'):
        self.cell.border = Border(left   = Side(border_style = border_style, color = border_color),
                                  right  = Side(border_style = border_style, color = border_color),
                                  top    = Side(border_style = border_style, color = border_color),
                                  bottom = Side(border_style = border_style, color = border_color)
                                  )

    def edit_border_left(self, border_style = 'thin', border_color = 'FF000000'):
        self.cell.border = Border(left = Side(border_style = border_style, color = border_color))
    
    def edit_border_right(self, border_style = 'thin', border_color = 'FF000000'):
        self.cell.border = Border(right = Side(border_style = border_style, color = border_color))

    def edit_border_top(self, border_style = 'thin', border_color = 'FF000000'):
        self.cell.border = Border(top = Side(border_style = border_style, color = border_color))

    def edit_border_bottom(self, border_style = 'thin', border_color = 'FF000000'):
        self.cell.border = Border(bottom = Side(border_style = border_style, color = border_color))
    # 文字の位置を整列させる
    def edit_alignment(self, horizontal = 'center', vertical = 'center'):
        self.cell.alignment = Alignment(horizontal = horizontal, vertical = vertical)

    # セルを塗りつぶしする
    def edit_fill(self, fill_type = 'solid', fgColor = 'FFFFFFFF'):
        self.cell.fill = PatternFill(fill_type = fill_type, fgColor = fgColor)
    
    def edit_height_width(self, width = 8.11, height = 13.2):
        self.ws.column_dimensions[get_column_letter(self.column)].width = width
        self.ws.row_dimensions[self.row].height = height

if __name__ == '__main__':
    sheet_name = 'test'
    filename = 'example.xlsx'
    work_book = openpyxl.Workbook()
    work_sheet = work_book.create_sheet(sheet_name)
    cell = EditCell(work_sheet, 2, 2)
    cell.edit_font(bold = True)
    cell.edit_border_left()
    cell.edit_alignment()
    work_book.save(filename)
