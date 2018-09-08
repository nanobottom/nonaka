import openpyxl, datetime, locale, calendar, jholiday, os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from editcell import EditCell
current_year = 2018
current_month = 4
fin_month_interval = 1
white_row_num = 30
main_column_num = 8
project_num = 3
filename = '作業進捗表.xlsx'

current_dir = os.path.dirname(os.path.abspath(__file__))
xlsx_path = os.path.join(current_dir, filename)
try:
    if os.path.exists(xlsx_path) == False:
        raise SystemError('Missing "{0}" file at {1}. Make new file.'.format(filename, current_dir))
except SystemError as e:
    print("SystemError : {}".format(e))
    work_book = openpyxl.Workbook()
else:
    work_book = load_workbook(filename)
sheet_name = str(current_year) + '年' + str(current_month) + '月'
work_sheet = work_book.create_sheet(sheet_name)

# 土日祝日を判定して土曜日なら青,祝日か日曜日なら赤を、それ以外ならcolorを返す関数
def is_sat_sun_holiday(date, color):
    royalblue = 'FF4169E1'
    tomato = 'FFFF6347'
    day_of_the_week_index = int(date.strftime('%w'))
    if jholiday.holiday_name(date.year, date.month, date.day) is not None:
        return tomato
    else:
        if day_of_the_week_index == 6:
            return royalblue
        elif day_of_the_week_index == 0:
            return tomato
        else:
            return color

# カラーコード
paleturquoise = 'FFAFEEEE'
ghostwhite    = 'FFF8F8FF'
darkorange    = 'FFFF8C00'
yellow        = 'FFFFD700'

# 番号(A1)のセル編集
work_sheet.merge_cells("A1:A3")
a1 = EditCell(work_sheet, 1, 1, '番号')
a1.edit_font(bold = True)
a1.edit_border_left()
a1.edit_border_top()
a1.edit_border_right()
a1.edit_height_width(width = 6)
a1.edit_fill(fgColor = paleturquoise)
a1.edit_alignment()

a2 = EditCell(work_sheet, 1, 2)
a2.edit_border_left()
a2.edit_border_right()

a3 = EditCell(work_sheet, 1, 3)
a3.edit_border_left()
a3.edit_border_right()
a3.edit_border_bottom()

# 項目(B1)のセル編集
work_sheet.merge_cells("B1:B3")
b1 = EditCell(work_sheet, 2, 1, '項目')
b1.edit_font(bold = True)
b1.edit_border_left()
b1.edit_border_top()
b1.edit_border_right()
b1.edit_height_width(width = 30)
b1.edit_fill(fgColor = paleturquoise)
b1.edit_alignment()

b2 = EditCell(work_sheet, 2, 2)
b2.edit_border_left()
b2.edit_border_right()

b3 = EditCell(work_sheet, 2, 3)
b3.edit_border_left()
b3.edit_border_right()
b3.edit_border_bottom()
# 予定(C1)のセル編集
work_sheet.merge_cells("C1:D2")
c1 = EditCell(work_sheet, 3, 1, '予定')
c1.edit_font(bold = True)
c1.edit_border_left()
c1.edit_border_top()
c1.edit_fill(fgColor = paleturquoise)
c1.edit_alignment()

c2 = EditCell(work_sheet, 3, 2)
c2.edit_border_left()
c2.edit_border_bottom()

d1 = EditCell(work_sheet, 4, 1)
d1.edit_border_right()
d1.edit_border_top()

d2 = EditCell(work_sheet, 4, 2)
d2.edit_border_right()
d2.edit_border_bottom()

# 実績(E1)のセル編集
work_sheet.merge_cells("E1:F2")
e1 = EditCell(work_sheet, 5, 1, '実績')
e1.edit_font(bold = True)
e1.edit_border_left()
e1.edit_border_top()
e1.edit_fill(fgColor = paleturquoise)
e1.edit_alignment()

e2 = EditCell(work_sheet, 5, 2)
e2.edit_border_left()
e2.edit_border_bottom()

f1 = EditCell(work_sheet, 6, 1)
f1.edit_border_right()
f1.edit_border_top()

f2 = EditCell(work_sheet, 6, 2)
f2.edit_border_right()
f2.edit_border_bottom()

#開始(C3)のセル編集
c3 = EditCell(work_sheet, 3, 3, '開始')
c3.edit_font(bold = True)
c3.edit_fill(fgColor = paleturquoise)
c3.edit_height_width(width = 6)
c3.edit_alignment()
c3.edit_border_round()

#終了(D3)のセル編集
d3 = EditCell(work_sheet, 4, 3, '終了')
d3.edit_font(bold = True)
d3.edit_fill(fgColor = paleturquoise)
d3.edit_height_width(width = 6)
d3.edit_alignment()
d3.edit_border_round()

#開始(E3)のセル編集
e3 = EditCell(work_sheet, 5, 3, '開始')
e3.edit_font(bold = True)
e3.edit_fill(fgColor = paleturquoise)
e3.edit_height_width(width = 6)
e3.edit_alignment()
e3.edit_border_round()

#終了(F3)のセル編集
f3 = EditCell(work_sheet, 6, 3, '終了')
f3.edit_font(bold = True)
f3.edit_fill(fgColor = paleturquoise)
f3.edit_height_width(width = 6)
f3.edit_alignment()
f3.edit_border_round()

#工数(G1)のセル編集
work_sheet.merge_cells("G1:G3")
g1 = EditCell(work_sheet, 7, 1, '工数')
g1.edit_font(bold = True)
g1.edit_border_left()
g1.edit_border_top()
g1.edit_border_right()
g1.edit_height_width(width = 6)
g1.edit_fill(fgColor = paleturquoise)
g1.edit_alignment()

g2 = EditCell(work_sheet, 7, 2)
g2.edit_border_left()
g2.edit_border_right()

g3 = EditCell(work_sheet, 7, 3)
g3.edit_border_left()
g3.edit_border_right()
g3.edit_border_bottom()


#状態(H1)のセル編集
work_sheet.merge_cells("H1:H3")
h1 = EditCell(work_sheet, 8, 1, '状態')
h1.edit_font(bold = True)
h1.edit_border_left()
h1.edit_border_top()
h1.edit_border_right()
h1.edit_fill(fgColor = paleturquoise)
h1.edit_alignment()

h2 = EditCell(work_sheet, 8, 2)
h2.edit_border_left()
h2.edit_border_right()

h3 = EditCell(work_sheet, 8, 3)
h3.edit_border_left()
h3.edit_border_right()
h3.edit_border_bottom()

#白枠（4行目以降）の編集
for i_row in range(4, white_row_num + 4):
    for i_column in range(1, main_column_num + 1):
        white = EditCell(work_sheet, i_column, i_row)
        white.edit_fill(fgColor = ghostwhite)
        white.edit_alignment()
        white.edit_border_round()

# 列の固定
work_sheet.freeze_panes = 'I1'

date = datetime.datetime(current_year, current_month, 1)
column_count = main_column_num + 1
for i_month in range(current_month, current_month + fin_month_interval):
    move_up_year = i_month // 12
    if move_up_year > 0 and i_month % 12 != 0:
        i_month -= move_up_year * 12
    (_, last_day) = calendar.monthrange(current_year + move_up_year, i_month)

    for i_day in range(1,last_day + 1):
        # 年月日を取得する
        year = date.year
        month = date.month
        day = date.day

        # 日付の文字列からその日付の曜日を取得する
        locale.setlocale(locale.LC_TIME, 'ja_JP.UTF-8')
        day_of_the_week = date.strftime('%a')
        day_of_the_week_index = int(date.strftime('%w'))
        
        # 日付セルの編集
        if i_day == 1:
            # 年度の最初の月はxxxx年xx月の文字列を表示させるためにセルを結合する
            if i_month == 1:
                month_title = str(year) + '年' + str(month) + '月'
                work_sheet.merge_cells(get_column_letter(column_count) + '1:' + get_column_letter(column_count + 2) + '1')
            # 年度の最初の月以外はxx月と表示させる
            else:
                work_sheet.merge_cells(get_column_letter(column_count) + '1:' + get_column_letter(column_count + 1) + '1')
                month_title = str(month) + '月'
            i1= EditCell(work_sheet, column_count, 1, month_title)
            i1.edit_border_left()
        else:
            i1= EditCell(work_sheet, column_count, 1)
        i1.edit_font(bold = True)
        i1.edit_fill(fgColor = darkorange)
        i1.edit_height_width(width = 4)
        i1.edit_alignment()

        #土曜日なら青、日曜日、祝日なら赤、それ以外は黄色にして日付と曜日を表示させる
        ans_color = is_sat_sun_holiday(date, yellow)
        i2= EditCell(work_sheet, column_count, 2, str(day))
        i2.edit_font(bold = True)
        i2.edit_fill(fgColor = ans_color)
        i2.edit_height_width(width = 4)
        i2.edit_alignment()
        i2.edit_border_round()

        i3 = EditCell(work_sheet, column_count, 3, day_of_the_week)
        i3.edit_font(bold = True)
        i3.edit_fill(fgColor = ans_color)
        i3.edit_height_width(width = 4)
        i3.edit_alignment()
        i3.edit_border_round()

        #白枠（4行目以降）の編集(土日祝で色を変更する)
        for i_row in range(4, white_row_num + 4):
            ans_color = is_sat_sun_holiday(date, ghostwhite)
            white_cell = EditCell(work_sheet, column_count, i_row)
            white_cell.edit_fill(fgColor = ans_color)
            white_cell.edit_height_width(width = 4)
            white_cell.edit_alignment()
            white_cell.edit_border_round()
        print('Processing : {0}年{1}月{2}日({3})'.format(year, month, day, day_of_the_week))
        date = date + datetime.timedelta(days = 1)
        column_count += 1
    current_month += 1

# 工数の数式（全日程の時間の合計を算出する）を代入する
for i_row in range(4,white_row_num + 4):
    numerical_formula_sum ='=SUM(I'+str(i_row)+':'+get_column_letter(column_count - 1)+str(i_row)+')'
    work_sheet.cell(row = i_row, column = 7).value = numerical_formula_sum

# プロジェクト毎の工数を算出する
for i_project in range(1, project_num + 1):
    numerical_formula_sumif = '=SUMIF(A4:A' + str(white_row_num + 4) + ',' + str(i_project) + ',G4:G' + str(white_row_num + 4) + ')'
    p_title_cell = EditCell(work_sheet, i_project + 2, white_row_num + 6, 'P' + str(i_project) )
    p_title_cell.edit_font(bold = True)
    p_title_cell.edit_fill(fgColor = paleturquoise)
    p_title_cell.edit_alignment()
    p_title_cell.edit_border_round()

    nf_cell = EditCell(work_sheet, i_project + 2, white_row_num + 7, numerical_formula_sumif)
    nf_cell.edit_font(bold = True)
    nf_cell.edit_fill(fgColor = ghostwhite)
    nf_cell.edit_alignment()
    nf_cell.edit_border_round()



work_book.save(filename)
print('Finish!')
