# -*- coding: utf-8 -*-
import pyautogui
import webbrowser
import time
import os
import openpyxl
from PIL import ImageGrab

#config.iniを読み込む

class FormAutomation:
    def __init__(self):
        pyautogui.PAUSE = 0.1
        pyautogui.FAILSAFE = True
        self.TARGET_URL = 'http://autbor.com/form'
        #self.TARGET_URL = ''
        self.SPREADSHEET_NAME = 'form_data.xlsx'
        self.WAIT_TIME = 5
        self.SCROLL_NUM = -200
        self.WRITE_SPEED = 0.5
        self.NAME_FIELD = (743, 367)
        self.SUBMIT_ANOTHER_LINK = (774, 271)
        self.form_datas = list()
    
    def open_target_form(self):
        if self.TARGET_URL:
            webbrowser.open(self.TARGET_URL)
        
    def resolution_check(self):
        # 画面の解像度が一致しているか判定
        width, height = pyautogui.size()
        if not (width == 1920 and height == 1080):
            raise Warning('モニターの解像度を1920 x 1080に設定してください。')
        # スクリーンショット用のフォルダを作成する
        os.makedirs('screenshot', exist_ok=True)
    
    def read_spreadsheet(self):
        wb = openpyxl.load_workbook(self.SPREADSHEET_NAME)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        title = list()
        for row_num in range(1, ws.max_row+1):
            form_data = dict()
            for column_num in range(1, ws.max_column+1):
                if row_num == 1:
                    title.append(ws.cell(row=row_num, column=column_num).value)
                else:
                    form_data[title[column_num-1]] = ws.cell(row=row_num, column=column_num).value
            if row_num != 1:
                self.form_datas.append(form_data)
        print(self.form_datas)
 
    def display_attention(self):
        # カーソルを左上に持っていくかctrl+alt+delでログアウトすればストップできることを表示する
        print('プログラムを止める場合はカーソルを左上に持っていくか、Ctrl+Alt+Delでログアウトすること。')
        # 入力フォームを全画面で表示するよう指示を出す
        print('{}秒以内に対象とするフォーム画面をWindows+↑キーで最大化してください。'.format(self.WAIT_TIME))
        time.sleep(self.WAIT_TIME)
        
    def input_form_by_image_recognition(self):
        # 画像認識は動作が遅く、クリックする場所を間違えるため、難あり
        # position_pictureからクリックする場所の画像を一覧として出す
        position_pictures = [filename for filename in os.listdir('position_picture')]
        print(position_pictures)
        for filename in position_pictures:
            # 画像認識で該当する場所を特定する。もしなければ下に少しスクロールさせてまた探す
            locate_screen = pyautogui.locateOnScreen(os.path.join('position_picture', filename))
            while not locate_screen:
                pyautogui.scroll(self.SCROLL_NUM)
                locate_screen = pyautogui.locateOnScreen(os.path.join('position_picture', filename))
                #raise Warning('クリックしたい画面が存在しません。')
                
            pyautogui.click(pyautogui.center(locate_screen))
        # 画像フォルダにスクリーンショットを撮って保存する    
    
    def input_form(self):
        for i, form_data in enumerate(self.form_datas):
            # ユーザがスクリプトを中断する機会を与える
            print('>>>5秒間停止中。中断するにはカーソルを左上に持っていってください。')
            # スクリーンショットを既に取っている場合はスルーする
            if True in [form_data['name'] in filename for filename in os.listdir('screenshot')]:
                pyautogui.alert('{}は既に処理が終わっています。'.format(form_data['name']))
                continue
            if i == 0:
                for i in range(5):
                    pyautogui.hotkey('ctrl', '-')
                
            # Name欄を入力する
            time.sleep(3)
            pyautogui.click(self.NAME_FIELD[0], self.NAME_FIELD[1])
            time.sleep(0.5)
            """
            locate_screen = pyautogui.locateOnScreen(os.path.join('position_picture', '001_name.png'))
            while not locate_screen:
                print('対象とする入力フォームが見つかりません。')
                time.sleep(10)
                locate_screen = pyautogui.locateOnScreen(os.path.join('position_picture', '001_name.png'))
            pyautogui.click(pyautogui.center(locate_screen))
            """
            pyautogui.typewrite(form_data['name'], interval=self.WRITE_SPEED)
            pyautogui.typewrite(['enter', '\t'], interval=self.WRITE_SPEED)
            time.sleep(0.5)
            # Greatest Fear(s)欄を入力する
            pyautogui.typewrite(form_data['fear'], interval=self.WRITE_SPEED)
            pyautogui.typewrite(['enter', '\t'], interval=self.WRITE_SPEED)
            # Source of Wizard Powers欄を入力する
            if form_data['source'] == 'wand':
                pyautogui.typewrite(['down', 'enter', '\t'], interval=self.WRITE_SPEED)
            elif form_data['source'] == 'amulet':
                pyautogui.typewrite(['down', 'down', 'enter', '\t'], interval=self.WRITE_SPEED)    
            elif form_data['source'] == 'crystal ball':
                pyautogui.typewrite(['down', 'down', 'down', 'enter', '\t'], interval=self.WRITE_SPEED)
            elif form_data['source'] == 'money':
                pyautogui.typewrite(['down', 'down', 'down', 'down', 'enter', '\t'], interval=self.WRITE_SPEED)
            # RoboCop欄を入力する
            if form_data['robocop'] == 1:
                pyautogui.typewrite([' ', '\t'], interval=self.WRITE_SPEED)
            elif form_data['robocop'] == 2:
                pyautogui.typewrite(['right', '\t'], interval=self.WRITE_SPEED)    
            elif form_data['robocop'] == 3:
                pyautogui.typewrite(['right', 'right', '\t'], interval=self.WRITE_SPEED)
            elif form_data['robocop'] == 4:
                pyautogui.typewrite(['right', 'right', 'right', '\t'], interval=self.WRITE_SPEED)
            elif form_data['robocop'] == 5:
                pyautogui.typewrite(['right', 'right', 'right', 'right', '\t'], interval=self.WRITE_SPEED)
            # Additional Comments欄を入力する
            time.sleep(0.5)
            pyautogui.typewrite(form_data['comments'], interval=self.WRITE_SPEED)
            pyautogui.typewrite(['enter' , '\t'], interval=self.WRITE_SPEED)
            # 画面全体を表示させ、スクリーンショットを撮る
            time.sleep(3)
            ImageGrab.grab().save(os.path.join('screenshot', form_data['name'] + '.png'))
            print('スクリーンショットを撮りました。')
            # Submitをクリックする
            pyautogui.press('enter')
            # 次のページが読み込まれるのを待つ
            print('送信ボタンを押しました')
            time.sleep(2)
            # Submit another responseリンクをクリックする
            pyautogui.click(self.SUBMIT_ANOTHER_LINK[0], self.SUBMIT_ANOTHER_LINK[1])
            
            
            
            
            


# スプレッドシートを読み込む
# スクショを既に撮っている場合は飛ばす
### 画像ファイル名とスプレッドシートで入力したい文字列のキー（辞書で管理）を対応づけておく
# 画像認識を使用して指定したフォルダの名前順(send.xxを除く)にテキストボックスに文字列を入力する

# 画像認識を使用してsend.xxのセンターをクリックしてフォームを送信する

# full screen
#ImageGrab.grab().save("PIL_capture.png")

# 指定した領域内をクリッピング
#ImageGrab.grab(bbox=(100, 100, 200, 200)).save("PIL_capture_clip.png")

if __name__ == '__main__':
    form_automation = FormAutomation()
    form_automation.open_target_form()
    form_automation.resolution_check()
    form_automation.read_spreadsheet()
    form_automation.display_attention()
    form_automation.input_form()
